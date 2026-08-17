import pandas as pd, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = {
 "A": r"C:\360极速浏览器下载\25.7.30-25.8.15.xlsx",
 "B": r"C:\360极速浏览器下载\销售主题分析_款_20260731091709_181259840_1.xlsx",
 "C": r"C:\360极速浏览器下载\销售主题分析_款_20260731091733_181260081_1.xlsx",
 "D": r"C:\360极速浏览器下载\销售主题分析_商品_20260731092116_181262652_1.xlsx",
}

def load(f, col):
    df = pd.read_excel(f)
    df['code'] = df[col].astype(str).str.strip()
    df = df[~df['code'].str.contains('#|总|计', na=False)]
    for c in ['销售数量', '销售金额']:
        df[c] = pd.to_numeric(df[c], errors='coerce')
    return df

A = load(PATH['A'], '款式编码/商品编码')
B = load(PATH['B'], '款式编码/商品编码')
C = load(PATH['C'], '款式编码/商品编码')
D = load(PATH['D'], '款式编码')

def dmap(df):
    return {r['code']: (r['销售数量'], r['销售金额'], r['产品分类']) for _, r in df.iterrows()}

Ad, Bd, Cd = dmap(A), dmap(B), dmap(C)
cat = {}
for d in (Bd, Ad, Cd):
    for k, v in d.items():
        cat.setdefault(k, v[2])

def csum(d, i):
    a = {}
    for k, v in d.items():
        c = cat.get(k, '未知')
        a[c] = a.get(c, 0) + v[i]
    return a

Bq = {k: v[0] for k, v in Bd.items()}
Aq = {k: v[0] for k, v in Ad.items()}
Cq = {k: v[0] for k, v in Cd.items()}
Bcat = csum(Bd, 0); Ccat = csum(Cd, 0); Acat = csum(Ad, 0)
K = 10

# ---- 款式级预估 ----
style_rows = []
for code in Bd:
    a = Ad.get(code, (0, 0, cat[code]))[0]
    b = Bd[code][0]
    c = Cd.get(code, (0, 0, cat[code]))[0]
    ct = cat[code]
    Bc = Bcat.get(ct, 0); Cc = Ccat.get(ct, 0)
    if a > 0:
        geff = (b + K * Bc) / (c + K * Cc) if (c + K * Cc) > 0 else 1.0
        est_q = round(a * geff, 1)
        meth = '款式同比(收缩)'
    else:
        rl = Acat.get(ct, 0) / Cc if Cc > 0 else (sum(Aq.values()) / sum(Cq.values()))
        est_q = round(b * rl, 1)
        meth = '新品·品类季节系数'
    bp = Bd[code][1]; bq = Bd[code][0]
    up = bp / bq if bq > 0 else 0
    est_a = round(est_q * up, 2)
    style_rows.append([code, ct, a, b, c, est_q, est_a, meth])

style_df = pd.DataFrame(style_rows, columns=['款式编码', '产品分类', 'A_25late数量', 'B_26early数量', 'C_25early数量', '预估数量', '预估金额', '预估方法'])
style_df = style_df.sort_values('预估数量', ascending=False).reset_index(drop=True)

# ---- SKC 拆分 (用D的颜色占比) ----
D['color'] = D['颜色规格'].astype(str).str.split(';').str[0].str.strip()
skc_rows = []
unsplit = []
for _, r in style_df.iterrows():
    code = r['款式编码']; est_q = r['预估数量']; est_a = r['预估金额']
    sub = D[D['code'] == code]
    if len(sub) == 0:
        unsplit.append([code, r['产品分类'], est_q, est_a, '未拆分(缺颜色明细)'])
        continue
    g = sub.groupby('color').agg(q=('销售数量', 'sum'), amt=('销售金额', 'sum'))
    tq = g['q'].sum(); ta = g['amt'].sum()
    for color, row in g.iterrows():
        qs = row['q'] / tq if tq > 0 else 0
        as_ = row['amt'] / ta if ta > 0 else 0
        skc_rows.append([code, color, r['产品分类'], round(est_q * qs, 1), round(est_a * as_, 2), round(qs * 100, 1), round(as_ * 100, 1)])

skc_df = pd.DataFrame(skc_rows, columns=['款式编码', '颜色', '产品分类', '预估数量', '预估金额', '数量占比%', '金额占比%'])
unsplit_df = pd.DataFrame(unsplit, columns=['款式编码', '产品分类', '预估数量', '预估金额', '拆分状态'])

print('款式级预估总量: 数量', round(style_df['预估数量'].sum()), ' 金额', round(style_df['预估金额'].sum()))
print('SKC已拆分款数:', skc_df['款式编码'].nunique(), ' SKC行数:', len(skc_df))
print('未拆分款数:', len(unsplit_df))
print('SKC数量合计(应≈款式级):', round(skc_df['预估数量'].sum() + unsplit_df['预估数量'].sum()))

# ---- 写Excel ----
wb = Workbook()
thin = Side(style='thin', color='D0D0D0')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill('solid', fgColor='4472C4')
hdr_font = Font(bold=True, color='FFFFFF')
title_font = Font(bold=True, size=13, color='1F3864')

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

def write_df(ws, df, start=1):
    for j, col in enumerate(df.columns, 1):
        ws.cell(row=start, column=j, value=col)
    style_header(ws, start, len(df.columns))
    for i, (_, rr) in enumerate(df.iterrows(), start + 1):
        for j, col in enumerate(df.columns, 1):
            v = rr[col]
            cell = ws.cell(row=i, column=j, value=(None if pd.isna(v) else v))
            cell.border = border
            if isinstance(v, (int, float)) and col not in ('款式编码', '颜色', '产品分类', '预估方法', '拆分状态'):
                if col in ('预估数量', '数量占比%'):
                    cell.number_format = '#,##0.0'
                else:
                    cell.number_format = '#,##0.00'
    for j, col in enumerate(df.columns, 1):
        if len(df):
            w = max(len(str(col)), max(len(str(rr[col])) for _, rr in df.iterrows()))
        else:
            w = len(str(col))
        ws.column_dimensions[get_column_letter(j)].width = min(max(w + 2, 10), 24)

ws1 = wb.active; ws1.title = 'SKC级预估'
ws1.cell(row=1, column=1, value='26年7.30-8.15 销售SKC预估（款式×颜色）').font = title_font
write_df(ws1, skc_df, start=3)
ws1.freeze_panes = 'A4'

ws2 = wb.create_sheet('款式级预估')
ws2.cell(row=1, column=1, value='26年7.30-8.15 销售款式级预估').font = title_font
write_df(ws2, style_df, start=3)
ws2.freeze_panes = 'A4'

ws3 = wb.create_sheet('未拆分款式')
ws3.cell(row=1, column=1, value='未拆分SKC的款式（SKU文件缺颜色明细）').font = title_font
write_df(ws3, unsplit_df, start=3)

ws4 = wb.create_sheet('说明与方法')
notes = [
    ('数据核对', 'A=25年7.30-8.15款式; B=26年7.1-7.29款式; C=25年7.1-7.29款式; D=商品/SKU(按用户指示作为26年7.1-7.29颜色来源)'),
    ('重要提示', '经交叉核对，D文件总量与C完全一致、且在共有款中100%匹配C；但按用户指示，已采用D的款式颜色占比用于SKC拆分。颜色占比为相对值，不受年份口径影响。'),
    ('SKC覆盖', 'D仅含B的74/410个款的颜色明细，其余336款无法拆分到颜色，列入「未拆分款式」表(按款式级预估值)。'),
    ('同比方法', '同周期同比：26late ≈ 25late × (26early/25early)。采用品类收缩估计 geff=(B_s+K·B_cat)/(C_s+K·C_cat)，K=10，避免25年早期基数极小时产生极端值。'),
    ('新品处理', '无25late基数(a=0)的款：预估 = B_26early × 品类季节系数(A_cat/C_cat)，向前推算。'),
    ('金额口径', '预估金额 = 预估数量 × 26年early款式单价(销售金额/销售数量)，采用26年当前定价。'),
    ('整体对照', 'B(26early)=3737件; C(25early)=2844件; 整体同比B/C=1.314; 整体季节A/C=1.296; 款式级预估总量≈7269件。'),
]
ws4.cell(row=1, column=1, value='说明与方法').font = title_font
for i, (k, v) in enumerate(notes, 3):
    ws4.cell(row=i, column=1, value=k).font = Font(bold=True)
    ws4.cell(row=i, column=2, value=v).alignment = Alignment(wrap_text=True, vertical='top')
    ws4.column_dimensions['A'].width = 14
    ws4.column_dimensions['B'].width = 110

out = r"C:\Users\Administrator\WorkBuddy\1688业务\26年7.30-8.15_SKC预估.xlsx"
wb.save(out)
print('\n已保存:', out)
