import pandas as pd, re, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

skc_text = ("XQ6CS427 粉紫 XQ6CS427 白色 XQ6DD507 深灰 XQ6DD507 黑色 XQ6DD507 藏青 "
"XQ6DD507 浅灰 XQ6KT551 粉色 XQ6KT551 杏色 XQ6KT553 粉色 XQ6KT553 花灰 "
"XQ6KT554 粉色 XQ6KT554 酒红 XQ6KT554 藏青色 XQ6MY441 花灰色 XQ6MY441 咖色 "
"XQ6MY441 黑色 XQ6MY441 藏青色 XQ6MY441 藏青 XQ6MY441 花灰 XQ6NZ533 蓝色 "
"XQ6QK510 浅灰 XQ6QK510 深紫 XQ6QK510 深灰 XQ6SK527 酒红色 XQ6SK527 藏青 "
"XQ6TX405 咖色 XQ6TX405 粉色 XQ6TX405 白色 XQ6TX405 红色 XQ6TX406 奶黄 "
"XQ6TX406 灰紫 XQ6TX406 浅粉 XQ6TX406 杏色 XQ6TX407 黄色 XQ6TX407 米色 "
"XQ6TX410 粉色 XQ6TX410 藏青 XQ6TX410 白色 XQ6TX414 藏青 XQ6TX414 杏色 "
"XQ6TX414 粉色 XQ6TX414 紫红 XQ6TX570 花灰 XQ6TX570 藏青 XQ6WT483 橘色 "
"XQ6WT483 浅紫 XQ6WT483 黄色 XQ6WT483 黑色 XQ6WT485 黄色 XQ6WT485 玫红 "
"XQ6WT485 藏青 XQ6WT487 黄色 XQ6WT487 蓝色 XQ6WT487 玫红 XQ6WT488 大红 "
"XQ6WT488 黄色 XQ6WY433 玫红 XQ6WY433 大红 XQ6WY433 深咖 XQ6WY433 紫色 "
"XQ6WY433 粉色 XQ6WY436 杏色 XQ6WY436 深灰 XQ6WY436 紫色 XQ6WY436 粉色")
pairs = re.findall(r'(XQ6\w+)\s+([^\s]+)', skc_text)

# 颜色归一化：统一去掉词尾"色"字（聚水潭里同一颜色存在"藏青"/"藏青色"、"酒红"/"酒红色"两种写法）
def norm_color(c):
    c = str(c).strip()
    return c[:-1] if (len(c) >= 2 and c.endswith('色')) else c

# 清单按归一化后的颜色去重（如 XQ6MY441 同时录了"花灰色"和"花灰"、"藏青色"和"藏青"）
style_color = {}
dedup_log = []
for code, color in pairs:
    lst = style_color.setdefault(code, [])
    if norm_color(color) in [norm_color(x) for x in lst]:
        dedup_log.append((code, color))
        continue
    lst.append(color)
codes = list(style_color.keys())

def load(f):
    d = pd.read_excel(f)
    d['款式编码'] = d['款式编码'].astype(str).str.strip()
    d = d[~d['款式编码'].str.contains('#|总|计', na=False)]
    d['颜色规格'] = d['颜色规格'].astype(str).str.strip()
    d['颜色'] = d['颜色规格'].str.split(';').str[0].str.strip()
    d['销售数量'] = pd.to_numeric(d['销售数量'], errors='coerce')
    d['销售金额'] = pd.to_numeric(d['销售金额'], errors='coerce')
    return d

jt26 = load(r"C:\360极速浏览器下载\销售主题分析_商品_20260810155006_183907082_1.xlsx")
jt25c = load(r"C:\360极速浏览器下载\销售主题分析_商品_20260810155110_183907502_1.xlsx")
jt25l = load(r"C:\360极速浏览器下载\销售主题分析_商品_20260810155151_183907820_1.xlsx")

# R_cat from 25 comparable
def agg_by_style(df):
    return df.groupby('款式编码').agg(销量=('销售数量','sum'), 分类=('产品分类','first')).reset_index()
a25c = agg_by_style(jt25c); a25l = agg_by_style(jt25l)
m25 = a25c.merge(a25l, on='款式编码', how='inner', suffixes=('_c','_l'))
m25 = m25[m25['销量_c']>0].rename(columns={'分类_c':'分类'})
m25['R'] = m25['销量_l']/m25['销量_c']
R_cat = m25.groupby('分类')['R'].median()
R_n = m25.groupby('分类')['R'].count()
R_all = float(m25['R'].median())

a26 = agg_by_style(jt26).set_index('款式编码')

# 颜色占比（JT_26 实际）
def color_shares(code):
    sub = jt26[jt26['款式编码']==code].copy()
    sub['颜色N'] = sub['颜色'].map(norm_color)
    g = sub.groupby('颜色N')['销售数量'].sum()
    tot = g.sum()
    return (g/tot).to_dict() if tot>0 else {}

style_rows=[]; skc_rows=[]; missing_colors=[]
for c in codes:
    s26 = float(a26.loc[c,'销量']) if c in a26.index else 0.0
    cat = str(a26.loc[c,'分类']) if c in a26.index else '未知'
    R = float(R_cat.get(cat, R_all)); nsample = int(R_n.get(cat,0))
    est_q = s26 * R
    up = jt26[jt26['款式编码']==c]
    q=up['销售数量'].sum(); am=up['销售金额'].sum()
    unit = float(am/q) if q>0 else 0.0
    est_a = est_q * unit
    style_rows.append([c, cat, s26, round(R,3), nsample, round(est_q,1), round(est_a,1)])
    # color split
    cs = color_shares(c)
    present={}; miss=[]
    for uc in style_color[c]:
        jc = norm_color(uc)
        if jc in cs: present[uc]=cs[jc]
        else: miss.append(uc)
    if miss:
        missing_colors.append((c, miss))
    # 缺色（近15天无销量：断货/刚上新）按同款已有颜色平均占比的 50% 估算，再整体归一
    if miss:
        base = (sum(present.values())/len(present)) if present else 1.0/len(miss)
        for uc in miss: present[uc] = base*0.5
    # renormalize listed colors
    tot = sum(present.values()) or 1
    for uc, sh in present.items():
        shn = sh/tot
        skc_rows.append([c, uc, cat, round(est_q*shn,1), round(est_a*shn,1), round(shn*100,1),
                         '近期无销量-占比估算' if uc in miss else 'JT近15天实际占比'])

style_df = pd.DataFrame(style_rows, columns=['款式编码','产品分类','S26近15天销量','R季节倍数','R样本数','预估数量','预估金额'])
skc_df = pd.DataFrame(skc_rows, columns=['款式编码','颜色','产品分类','预估数量','预估金额','颜色占比%','占比来源'])

print("=== 款式级预估 ===")
print(style_df.to_string(index=False))
print("\n预估总量: 数量", round(style_df['预估数量'].sum()), " 金额", round(style_df['预估金额'].sum()))
print("SKC行数:", len(skc_df), " SKC数量合计:", round(skc_df['预估数量'].sum(),1))
print("\n缺失颜色(近期无销量):", missing_colors)
print("清单去重(同色不同写法):", dedup_log)
chk = skc_df.groupby('款式编码')['颜色占比%'].sum().round(1)
print("占比合计异常款:", chk[(chk<99.5)|(chk>100.5)].to_dict())

# ---- 写Excel ----
wb = Workbook()
thin = Side(style='thin',color='D0D0D0'); border = Border(left=thin,right=thin,top=thin,bottom=thin)
hdr_fill = PatternFill('solid',fgColor='4472C4'); hdr_font = Font(bold=True,color='FFFFFF')
title_font = Font(bold=True,size=13,color='1F3864')
def style_header(ws,row,ncol):
    for cc in range(1,ncol+1):
        cell=ws.cell(row=row,column=cc); cell.fill=hdr_fill; cell.font=hdr_font
        cell.alignment=Alignment(horizontal='center',vertical='center'); cell.border=border
def write_df(ws,df,start=1):
    for j,col in enumerate(df.columns,1): ws.cell(row=start,column=j,value=col)
    style_header(ws,start,len(df.columns))
    for i,(_,rr) in enumerate(df.iterrows(),start+1):
        for j,col in enumerate(df.columns,1):
            v=rr[col]
            cell=ws.cell(row=i,column=j,value=(None if pd.isna(v) else v)); cell.border=border
            if isinstance(v,(int,float)) and col not in ('款式编码','颜色','产品分类','占比来源'):
                cell.number_format='#,##0.0' if '占比' in col or '销量' in col or '数量' in col else '#,##0.00' if '金额' in col else '#,##0.000'
    for j,col in enumerate(df.columns,1):
        w=max(len(str(col)),*(len(str(rr[col])) for _,rr in df.iterrows())) if len(df) else len(str(col))
        ws.column_dimensions[get_column_letter(j)].width=min(max(w+2,9),22)

ws1=wb.active; ws1.title='SKC级预估'
ws1.cell(row=1,column=1,value='26年8.10-9.10 销售SKC预估（款式×颜色）— Method2 聚水潭').font=title_font
write_df(ws1,skc_df,start=3); ws1.freeze_panes='A4'
ws2=wb.create_sheet('款式级预估')
ws2.cell(row=1,column=1,value='26年8.10-9.10 款式级预估（21款）').font=title_font
write_df(ws2,style_df,start=3); ws2.freeze_panes='A4'
ws3=wb.create_sheet('方法说明')
notes=[
 ('方法','Method2（聚水潭）：预估数量 = 26近15天销量 × 同类目25年季节倍数中位数 R。R = 25年8.10-9.10销量 / 25年同周期销量（同品类可比款中位数）。'),
 ('颜色拆分','按JT_26近15天各颜色实际销量占比拆分到SKC；标注"JT近15天实际占比"。颜色做了归一化（统一去掉词尾"色"字），因为聚水潭同一颜色存在"藏青"/"藏青色"、"酒红"/"酒红色"两种写法。'),
 ('清单去重','XQ6MY441 清单中"花灰色/花灰"、"藏青色/藏青"为同色重复录入，已合并，故SKC行数为63而非65。'),
 ('缺失颜色处理','清单颜色在JT_26近15天无销量时（断货或刚上新），按同款已有颜色平均占比的50%估算份额，再整体归一到100%，并标注"近期无销量-占比估算"。'),
 ('整体参数', f'R_all(全品类中位数)={round(R_all,3)}；21款在25年文件中重叠=0（均为26新款），故用同品类中位数而非款式自身倍数。'),
 ('重要提示-Method1未启用','生意参谋(sycm)的链接→款号映射在飞书多维表中无法落到21个XQ6新款：飞书2600条映射款号全为XQ5(25年)旧款，XQ6(26新款)为0；且136个sycm链接仅19个能匹配且均为25年款。故Method1(访客×转化×加购)暂无法逐款归因，未纳入本次预估。'),
 ('数据假设','JT_26近15天≈15天、JT_25年8.10-9.10=32天；R已对周期长度归一（S25l/S25c直接套用）。聚水潭单价采用26近15天实际均价。'),
 ('风险提示','26近15天基数极低(21款合计约560件)，R放大后敏感；部分类目25样本少(如裙裤n=2、打底裤样本少)倍数不稳，建议结合实际库存/上新节奏人工复核。'),
]
ws3.cell(row=1,column=1,value='方法说明与口径').font=title_font
for i,(k,v) in enumerate(notes,3):
    ws3.cell(row=i,column=1,value=k).font=Font(bold=True)
    ws3.cell(row=i,column=2,value=v).alignment=Alignment(wrap_text=True,vertical='top')
    ws3.column_dimensions['A'].width=16; ws3.column_dimensions['B'].width=120
# R_cat sheet
ws4=wb.create_sheet('各类目R倍数')
rc=pd.DataFrame({'产品分类':R_cat.index,'R季节倍数':R_cat.values,'25年样本数':R_n.values})
rc=rc.sort_values('R季节倍数',ascending=False)
write_df(ws4,rc,start=1)

out=r"C:\Users\Administrator\WorkBuddy\1688业务\26年8.10-9.10_XQ6新款_SKC预估.xlsx"
wb.save(out)
print("\n已保存:", out)
