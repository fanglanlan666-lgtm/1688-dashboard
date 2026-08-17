import pandas as pd, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NB = r"C:\360极速浏览器下载\销售主题分析_款_20260731100225_181285979_1.xlsx"  # 用户视为26early
ND = r"C:\360极速浏览器下载\销售主题分析_商品_20260731100238_181286089_1.xlsx"  # 用户视为26early SKU(标26秋)
A  = r"C:\360极速浏览器下载\25.7.30-25.8.15.xlsx"                              # 25late
C  = r"C:\360极速浏览器下载\销售主题分析_款_20260731091733_181260081_1.xlsx"   # 25early

def load(f, col):
    d = pd.read_excel(f); d['code'] = d[col].astype(str).str.strip()
    d = d[~d['code'].str.contains('#|总|计', na=False)]
    d['销售数量'] = pd.to_numeric(d['销售数量'], errors='coerce')
    d['销售金额'] = pd.to_numeric(d['销售金额'], errors='coerce')
    return d

NBd = load(NB,'款式编码/商品编码'); NDd = load(ND,'款式编码')
Ad  = load(A ,'款式编码/商品编码'); Cd  = load(C ,'款式编码/商品编码')

# 品类级 late/early 季节系数 r_cat = A_cat / C_cat（全量同类目）
def by_cat(df):
    g = df.groupby('产品分类')['销售数量'].sum()
    return g
Ac = by_cat(Ad); Cc = by_cat(Cd)
r_all = Ad['销售数量'].sum()/Cd['销售数量'].sum()
# 秋款整体季节系数(回退用)
aut = Ad[Ad['虚拟分类'].astype(str).str.contains('秋',na=False)]['销售数量'].sum()
cut = Cd[Cd['虚拟分类'].astype(str).str.contains('秋',na=False)]['销售数量'].sum()
r_aut = aut/cut if cut>0 else r_all

def r_cat(cat):
    a = Ac.get(cat,0); c = Cc.get(cat,0)
    if c >= 10 and a > 0:
        return a/c
    return r_aut   # 样本不足回退秋款整体

# 清单
raw = """XQ6QK510	浅灰
XQ6ZK520	黑色
XQ6QK510	深紫
XQ6QK510	深灰
XQ6ZK520	粉色
XQ6CS427	粉紫
XQ6ZK516	深灰
XQ6ZK516	浅粉
XQ6ZK516	紫红
XQ6DQ503	藏青
XQ6ZK516	花灰
XQ6CS427	白色
XQ6DQ503	咖色
XQ6DD507	深灰
XQ6DD507	黑色
XQ6DD507	藏青
XQ6MJ463	粉色
XQ6MJ463	玫紫
XQ6MJ463	黄绿
XQ6DD507	浅灰
XQ6TX422	浅卡其
XQ6TX422	杏色
XQ6NZ533	蓝色
XQ6TX422	粉色
XQ6SK526	深灰
XQ6TX405	咖色
XQ6TX405	粉色
XQ6KT551	粉色
XQ6WT487	黄色
XQ6WT487	蓝色
XQ6WT487	玫红
XQ6TX405	白色
XQ6QK563	粉色
XQ6KT551	杏色
XQ6QK563	黑白
XQ6SK526	赤红
XQ6TX405	红色
XQ6WT473	花灰
XQ6TX407	黄色
XQ6TX402	米色
XQ6WT488	大红
XQ6WY439	米色
XQ6TX402	黄色
XQ6WT488	黄色
XQ6TX407	米色
XQ6ZK520	深灰
XQ6WY439	粉色
XQ6TX402	紫色
XQ6TX402	嫩绿
XQ6ZK511	藏青
XQ6ZK511	咖色
XQ6ZK511	深灰
XQ6WY436	杏色
XQ6WY436	深灰
XQ6WY436	紫色
XQ6WY436	粉色
XQ6WY433	玫红
XQ6WY433	大红
XQ6WY433	深咖
XQ6WY433	紫色
XQ6WY433	粉色
XQ6WT485	黄色
XQ6WT485	玫红
XQ6WT485	藏青
XQ6WT476	紫色
XQ6WT476	藏青色
XQ6TX570	花灰
XQ6TX570	藏青
XQ6TX430	深灰
XQ6TX430	米色
XQ6TX430	粉色
XQ6TX416	白色
XQ6TX414	藏青
XQ6TX414	杏色
XQ6TX414	粉色
XQ6TX409	黄色
XQ6TX409	蓝色
XQ6TX406	奶黄
XQ6TX406	灰紫
XQ6TX406	浅粉
XQ6TX406	杏色
XQ6SK527	酒红色
XQ6SK527	藏青
XQ6SK523	绿色
XQ6SK523	粉色
XQ6SK523	紫色
XQ6SK522	粉色
XQ6SK522	酒红色
XQ6SK522	杏色
XQ6NZ538	深蓝
XQ6NZ538	蓝色
XQ6NZ537	蓝色
XQ6NZ537	藏青
XQ6MY441	花灰色
XQ6MY441	咖色
XQ6MY441	黑色
XQ6MY441	藏青色
XQ6MY441	藏青
XQ6MY441	花灰
XQ6MJ466	黑色
XQ6LY542	咖色
XQ6LY542	粉色
XQ6LY542	杏色
XQ6KT553	粉色
XQ6KT553	花灰
XQ6DQ501	紫色
XQ6DQ501	杏色
XQ6DQ501	浅蓝
XQ6DQ501	粉色
XQ6CS429	绿色
XQ6CS429	粉色
XQ6CS429	黄色
XQ6CS426	嫩绿
XQ6CS426	米色"""
pairs = [tuple(l.split('\t')) for l in raw.strip().splitlines()]
dfp = pd.DataFrame(pairs, columns=['款式编码','颜色'])

# 款式级：26early基准 = NB中该款销量；乘 r_cat
rows=[]
for code in sorted(dfp['款式编码'].unique()):
    b = pd.to_numeric(NBd[NBd['code']==code]['销售数量'],errors='coerce').sum()
    cat = NBd[NBd['code']==code]['产品分类']
    cat = cat.dropna().iloc[0] if len(cat.dropna()) else (Cd[Cd['code']==code]['产品分类'].dropna().iloc[0] if len(Cd[Cd['code']==code]['产品分类'].dropna()) else '未知')
    rc = r_cat(cat)
    est = b*rc
    amt = pd.to_numeric(NDd[NDd['code']==code]['销售金额'],errors='coerce').sum()
    qty = pd.to_numeric(NDd[NDd['code']==code]['销售数量'],errors='coerce').sum()
    unit = amt/qty if qty>0 else 0
    rows.append([code,cat,b,round(rc,3),round(est,1),round(est*unit,2)])
style_df = pd.DataFrame(rows, columns=['款式编码','产品分类','26early基准销量','类目季节系数','预估数量','预估金额'])
style_df = style_df.sort_values('预估数量',ascending=False).reset_index(drop=True)

# SKC拆分：用ND颜色占比（仅基于清单颜色相对归一化）
NDd['color'] = NDd['颜色规格'].astype(str).str.split(';').str[0].str.strip()
skc=[]
unmatched=[]
for code, sub in dfp.groupby('款式编码'):
    est_q = style_df.loc[style_df['款式编码']==code,'预估数量'].iloc[0]
    est_a = style_df.loc[style_df['款式编码']==code,'预估金额'].iloc[0]
    colors = list(sub['颜色'])
    # ND中该款颜色销量
    nd_sub = NDd[NDd['code']==code]
    w = {c: pd.to_numeric(nd_sub[nd_sub['color']==c]['销售数量'],errors='coerce').sum() for c in colors}
    tot = sum(w.values())
    if tot > 0:
        for c in colors:
            share = w[c]/tot
            skc.append([code,c,round(est_q*share,1),round(est_a*(share if est_a>0 else 0),2) if est_a>0 else 0, round(share*100,1)])
    else:
        # 该款在ND无颜色数据 → 清单颜色平均分配
        sh = 1.0/len(colors)
        for c in colors:
            skc.append([code,c,round(est_q*sh,1), round(est_a*sh,2), round(sh*100,1)])
            unmatched.append(code)
skc_df = pd.DataFrame(skc, columns=['款式编码','颜色','预估数量','预估金额','数量占比%'])

print("=== 清单款 26年7.30-8.15 款式级预估 (类目季节系数法) ===")
print("款式级预估总量: 数量", round(style_df['预估数量'].sum()), " 金额", round(style_df['预估金额'].sum()))
print("对照: 26early基准(NB)合计", round(style_df['26early基准销量'].sum()), " 秋款整体季节系数", round(r_aut,3), " 全量季节系数", round(r_all,3))
print("SKC行数", len(skc_df), " SKC数量合计", round(skc_df['预估数量'].sum()))
print("ND无颜色数据的款(已平均):", sorted(set(unmatched)))
print("\nTop15款式:")
print(style_df.head(15).to_string(index=False))

# 写Excel
wb = Workbook()
thin=Side(style='thin',color='D0D0D0'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
hf=PatternFill('solid',fgColor='4472C4'); hft=Font(bold=True,color='FFFFFF'); tf=Font(bold=True,size=13,color='1F3864')
def shead(ws,row,n):
    for c in range(1,n+1):
        cell=ws.cell(row=row,column=c); cell.fill=hf; cell.font=hft; cell.alignment=Alignment(horizontal='center',vertical='center'); cell.border=border
def wdf(ws,df,start=3):
    for j,col in enumerate(df.columns,1): ws.cell(row=start,column=j,value=col)
    shead(ws,start,len(df.columns))
    for i,(_,rr) in enumerate(df.iterrows(),start+1):
        for j,col in enumerate(df.columns,1):
            v=rr[col]; cell=ws.cell(row=i,column=j,value=(None if pd.isna(v) else v)); cell.border=border
            if isinstance(v,(int,float)) and col not in ('款式编码','颜色','产品分类'):
                cell.number_format='#,##0.0' if '占比' in col else '#,##0.00' if '金额' in col else '#,##0'
    for j,col in enumerate(df.columns,1):
        w=max(len(str(col)),*(len(str(rr[col])) for _,rr in df.iterrows())) if len(df) else len(str(col))
        ws.column_dimensions[get_column_letter(j)].width=min(max(w+2,11),22)
ws1=wb.active; ws1.title='SKC级预估'; ws1.cell(row=1,column=1,value='清单款 26年7.30-8.15 销售SKC预估(款×颜色)').font=tf
wdf(ws1,skc_df); ws1.freeze_panes='A4'
ws2=wb.create_sheet('款式级预估'); ws2.cell(row=1,column=1,value='清单款 26年7.30-8.15 款式级预估').font=tf
wdf(ws2,style_df); ws2.freeze_panes='A4'
ws4=wb.create_sheet('说明与方法')
notes=[
 ('方法逻辑','清单款为26秋新款，25年late无数据。按用户指示：26late ≈ 26early清单款销量 × 同类目(25late/25early)季节系数。'),
 ('26early基准','取自用户提供的NB文件(视作26年7.1-7.29)，清单款合计271件。经核对NB与25年early(C)逐款一致，此处直接采用NB销量作为26early基准。'),
 ('类目季节系数','r_cat = 该款产品分类在全量的 25late/25early 销量比(A/C)；样本不足(C<10或A=0)回退秋款整体系数=%.3f。' % r_aut),
 ('整体对照','秋款整体季节系数=%.3f；全量(所有品类)季节系数=%.3f；清单款整体预估=%.0f件。' % (r_aut,r_all,style_df['预估数量'].sum())),
 ('SKC颜色拆分','用ND(标2026;秋)的颜色占比，在清单指定颜色间相对归一化分配款式级预估；ND无颜色数据的款按清单颜色平均分配。'),
 ('金额口径','预估金额=预估数量 × ND中该款单价(销售金额/销售数量)；ND经核对实为25年同期数据，单价按此水平折算。'),
 ('无数据款','C=0且无26early基准的款(XQ6DQ503/XQ6LY542/XQ6SK526/XQ6WT476/XQ6WT485)预估为0，需补26early实际数据。'),
 ('重要提示','NB/ND经逐款比对与25年数据完全一致，实为25年同期数据被标为26年；本预估按用户口径以NB为26early基准执行。'),
]
ws4.cell(row=1,column=1,value='说明与方法').font=tf
for i,(k,v) in enumerate(notes,3):
    ws4.cell(row=i,column=1,value=k).font=Font(bold=True)
    ws4.cell(row=i,column=2,value=v).alignment=Alignment(wrap_text=True,vertical='top')
    ws4.column_dimensions['A'].width=16; ws4.column_dimensions['B'].width=110
out=r"C:\Users\Administrator\WorkBuddy\1688业务\清单款_26年7.30-8.15_SKC预估.xlsx"
wb.save(out)
print("\n已保存:", out)
PY