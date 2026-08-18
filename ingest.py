#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
1688 数字营销 · 半自动数据管道
把后台导出的 3 类 xlsx（总览 / 计划 / 商品）自动归一化、合并去重，
输出 master_data.csv 与 data.js（供 1688数字营销工作台.html 直接加载）。

用法：
  python ingest.py                      # 扫描 ./导出 目录下所有 xlsx
  python ingest.py "文件1.xlsx" "文件2.xlsx" ...   # 指定文件
输出：
  master_data.csv  统一长表（UTF-8 BOM）
  data.js          window.DASHBOARD_DATA = [...]
"""
import os, re, sys, glob, json
import openpyxl

WS = os.path.dirname(os.path.abspath(__file__))
OUT_CSV = os.path.join(WS, "master_data.csv")
OUT_JS  = os.path.join(WS, "data.js")
EXPORT_DIR = os.path.join(WS, "导出")
# 商品ID -> 款号 映射表（放在工作区内，保证管道可复现）
MAPPING_FILE = os.path.join(WS, "映射", "商品id款号映射.xlsx")
# 商品主图表（含 产品ID / 产品主图1），用于给商品维度关联主图
PROD_IMG_XLSX = os.path.join(WS, "产品主图.xlsx")
IMG_DIR = os.path.join(WS, "images")
# 预算配置文件（由看板「导出预算配置」生成，放在工作区根目录；ingest 合并进 data.js 实现永久保存）
BUDGET_FILE = os.path.join(WS, "budget.json")

# 源表头(中文) -> 规范字段
HEADER_MAP = {
    "展现": "imp", "曝光量": "imp",
    "有效曝光": "vexp",
    "点击": "clk", "点击量": "clk",
    "消耗": "cost", "消耗量": "cost",
    "点击率": "ctr", "点击转化率": "ctr",
    "平均点击成本": "cpc",
    "收藏商品数": "favP", "收藏店铺数": "favS", "加购数": "cart",
    "领券数": "coupon", "获取优惠券数": "coupon",
    "订单数": "ord", "广告引导提交订单数": "ord", "提交订单数": "ord",
    "询盘数": "inq",
    "线索量": "lead", "线索数": "lead",
    "成交金额": "gmv", "广告引导交易金额": "gmv",
    "广告投入产出比": "roi",
    "线索成本": "leadCost", "询盘成本": "inqCost",
    "线索转化率": "leadCvr", "询盘转化率": "inqCvr",
    "统计日期": "date",
    "一级产品": "plan",
    "商品ID": "pid", "商品标题": "title",
}
CANON = ["date","dim","plan","pid","title","sku","img","imp","vexp","clk","cost","ctr","cpc",
         "favP","favS","cart","coupon","ord","ordAd","inq","lead","gmv","roi","deal",
         "leadCost","inqCost","leadCvr","inqCvr"]

# 计划名称归一化（业务口径：计划维度的「解决方案」即商品维度的「大客方案」）
PLAN_ALIAS = {"解决方案": "大客方案"}

def to_float(v):
    if v is None: return None
    s = str(v).strip().replace(",", "")
    if s in ("", "—", "-", "None", "nan"): return None
    if s.endswith("%"):
        try: return float(s[:-1]) / 100.0
        except: return None
    try: return float(s)
    except: return None

def norm_date(v, fname=""):
    if v is None: v = ""
    s = str(v).strip()
    m = re.search(r"(\d{4})[-/]?(\d{2})[-/]?(\d{2})", s)
    if m: return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    # 文件名兜底
    fm = re.search(r"(\d{4}-\d{2}-\d{2})", fname)
    if fm: return fm.group(1)
    return ""

def detect_type(headers):
    h = set(headers)
    if "商品ID" in h:
        return "商品"
    if "一级产品" in h:
        return "计划"
    if "统计日期" in h and ("展现" in h or "曝光量" in h or "消耗" in h):
        return "总览"
    return None

def load_mapping():
    """加载 商品ID->款号 映射表。找不到则返回空字典（商品维度仅显示商品ID）。"""
    m = {}
    if not os.path.exists(MAPPING_FILE):
        print("⚠️ 未找到映射表 映射/商品id款号映射.xlsx，商品维度将仅显示商品ID。")
        return m
    try:
        wb = openpyxl.load_workbook(MAPPING_FILE, data_only=True)
    except Exception as e:
        print("⚠️ 映射表读取失败:", e); return m
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # 定位表头行
    hi = 0
    for i, row in enumerate(rows[:5]):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if "商品ID" in cells and "款号" in cells:
            hi = i; break
    headers = [str(c).strip() if c is not None else "" for c in rows[hi]]
    try:
        idi = headers.index("商品ID"); ski = headers.index("款号")
    except ValueError:
        print("⚠️ 映射表缺少 商品ID/款号 列。"); wb.close(); return m
    for row in rows[hi+1:]:
        if not row or row[idi] is None: continue
        pid = str(row[idi]).strip()
        sku = str(row[ski]).strip() if row[ski] is not None else ""
        if pid:
            m[pid] = sku
    wb.close()
    print(f"  加载映射表: {len(m)} 条 (商品ID→款号)")
    return m

def load_img_map():
    """加载 产品ID -> 产品主图1(URL)。找不到则返回空字典（商品维度不显示主图）。"""
    m = {}
    if not os.path.exists(PROD_IMG_XLSX):
        return m
    try:
        wb = openpyxl.load_workbook(PROD_IMG_XLSX, data_only=True)
    except Exception as e:
        print("⚠️ 主图表读取失败:", e); return m
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hi = 0
    for i, row in enumerate(rows[:5]):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if "产品ID" in cells and ("产品主图1" in cells or "主图1" in cells):
            hi = i; break
    headers = [str(c).strip() if c is not None else "" for c in rows[hi]]
    try: idi = headers.index("产品ID")
    except ValueError: wb.close(); return m
    try: imgi = headers.index("产品主图1")
    except ValueError:
        try: imgi = headers.index("主图1")
        except ValueError: wb.close(); return m
    for row in rows[hi+1:]:
        if not row or row[idi] is None: continue
        pid = str(row[idi]).strip()
        url = str(row[imgi]).strip() if row[imgi] else ""
        if pid and url.startswith("http"):
            m[pid] = url
    wb.close()
    print(f"  加载主图表: {len(m)} 条 (产品ID→主图URL)")
    return m

def download_images(pid_set, img_map):
    """把商品主图下载到 images/ 目录（已存在则跳过）。返回 pid->相对路径(images/xxx.jpg)。"""
    if not img_map: return {}
    os.makedirs(IMG_DIR, exist_ok=True)
    import urllib.request, ssl
    ctx = ssl.create_default_context()
    ctx.check_hostname = False; ctx.verify_mode = ssl.CERT_NONE
    result = {}; n = 0
    for pid in pid_set:
        url = img_map.get(pid)
        if not url:
            result[pid] = ""; continue
        dest = os.path.join(IMG_DIR, pid + ".jpg")
        if os.path.exists(dest) and os.path.getsize(dest) > 500:
            result[pid] = "images/" + pid + ".jpg"; continue
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            data = urllib.request.urlopen(req, timeout=20, context=ctx).read()
            if len(data) < 500:
                result[pid] = ""; continue
            with open(dest, "wb") as f: f.write(data)
            result[pid] = "images/" + pid + ".jpg"; n += 1
        except Exception as e:
            result[pid] = ""
    print(f"  主图下载: 新增 {n} 张，缓存目录 images/")
    return result

def read_sheet(ws, fname, mapping=None):
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    # 找表头行：含最多已知映射列的那行
    best, bidx, bscore = None, 0, -1
    for i, row in enumerate(rows[:10]):
        score = sum(1 for c in row if c is not None and str(c).strip() in HEADER_MAP)
        if score > bscore:
            best, bidx, bscore = row, i, score
    if best is None or bscore == 0: return []
    headers = [ (str(c).strip() if c is not None else "") for c in best ]
    idx = {HEADER_MAP[h]: ci for ci, h in enumerate(headers) if h in HEADER_MAP}
    dtype = detect_type(headers)
    if dtype is None: return []
    out = []
    for row in rows[bidx+1:]:
        if all(c is None or str(c).strip() == "" for c in row): continue
        def get(f):
            ci = idx.get(f)
            return row[ci] if (ci is not None and ci < len(row)) else None
        date = norm_date(get("date"), fname) if "date" in idx else norm_date(None, fname)
        if not date:
            print(f"  ⚠️ 跳过 {os.path.basename(fname)}：无法从表头/文件名识别日期。"
                  f"商品/总览表需文件名含 YYYY-MM-DD（如 *_2026-08-06-2026-08-06.xlsx），或表内含「统计日期」列。")
            return []
        rec = {k: None for k in CANON}
        rec["date"] = date
        rec["dim"] = dtype
        if dtype == "总览":
            rec["plan"] = "全部"
        elif dtype == "计划":
            plan = get("plan")
            if plan is None or str(plan).strip() in ("汇总", ""):
                continue  # 丢弃汇总行，避免与总览重复
            rec["plan"] = str(plan).strip()
        elif dtype == "商品":
            pid = get("pid"); title = get("title")
            if pid is None or str(pid).strip() in ("汇总", "", "-"):
                continue  # 丢弃汇总行
            rec["plan"] = "大客方案"   # 按用户定义：SKA驾驶舱=核心计划大客方案
            rec["pid"] = str(pid).strip()
            rec["sku"] = (mapping or {}).get(rec["pid"], "")  # 商品ID -> 款号
            rec["title"] = str(title).strip() if title else ""
        for f in ["imp","vexp","clk","cost","ctr","cpc","favP","favS","cart",
                  "coupon","ord","inq","lead","gmv","roi","leadCost","inqCost","leadCvr","inqCvr"]:
            if f in idx:
                rec[f] = to_float(get(f))
        # 派生指标
        if rec["ctr"] is None and rec["imp"] and rec["clk"]:
            rec["ctr"] = rec["clk"] / rec["imp"]
        if rec["cpc"] is None and rec["clk"] and rec["cost"]:
            rec["cpc"] = rec["cost"] / rec["clk"]
        if rec["roi"] is None and rec["cost"] and rec["gmv"]:
            rec["roi"] = rec["gmv"] / rec["cost"]
        out.append(rec)
    return out

def load_existing():
    if not os.path.exists(OUT_CSV): return {}
    existing = {}
    import csv
    with open(OUT_CSV, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            key = (r.get("date"), r.get("dim"), r.get("plan"), r.get("pid"))
            existing[key] = r
    return existing

def main():
    if len(sys.argv) > 1:
        files = sys.argv[1:]
    else:
        os.makedirs(EXPORT_DIR, exist_ok=True)
        files = glob.glob(os.path.join(EXPORT_DIR, "*.xlsx"))
    if not files:
        print("未找到 xlsx 文件，仅对已有 master_data.csv 做归一化与重写。")
    MAPPING = load_mapping()
    collected = []
    for fp in files:
        if not os.path.exists(fp):
            print("跳过(不存在):", fp); continue
        try:
            wb = openpyxl.load_workbook(fp, data_only=True)
        except Exception as e:
            print("打开失败:", fp, e); continue
        for ws in wb.worksheets:
            recs = read_sheet(ws, fp, MAPPING)
            if recs:
                print(f"  {os.path.basename(fp)} / {ws.title}: {len(recs)} 行 ({recs[0]['dim']})")
                collected.extend(recs)
        wb.close()
    # 合并去重（与已有 master 合并，相同 key 后者覆盖）
    store = load_existing()
    if not collected and not store:
        print("没有解析到任何数据行，且 master_data.csv 为空。"); return
    for r in collected:
        key = (r["date"], r["dim"], r["plan"], r["pid"] or "")
        store[key] = {k: ("" if r.get(k) is None else r.get(k)) for k in CANON}
    # 补齐所有记录的 CANON 字段（兼容旧 master 缺少的字段，如 img）
    for rec in store.values():
        for k in CANON:
            if k not in rec:
                rec[k] = ""
    merged = list(store.values())
    # 计划名称归一化（解决方案 → 大客方案），保证计划维度与商品维度口径一致
    for r in merged:
        if r.get("plan") in PLAN_ALIAS:
            r["plan"] = PLAN_ALIAS[r["plan"]]
    merged.sort(key=lambda x: (x["date"], x["dim"], x["plan"], x["pid"]))
    # 关联商品主图（按 产品ID 匹配并下载缓存）
    img_map = load_img_map()
    if img_map:
        pid_set = set(r.get("pid") for r in merged if r.get("dim") == "商品" and r.get("pid"))
        img_paths = download_images(pid_set, img_map)
        for r in merged:
            if r.get("dim") == "商品":
                r["img"] = img_paths.get(r.get("pid"), "")
    else:
        print("  （未找到 产品主图.xlsx，商品主图将不显示。把主图Excel放到工作区根目录命名为 产品主图.xlsx 即可。）")
    # 写 CSV
    import csv
    with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CANON)
        w.writeheader()
        for r in merged: w.writerow(r)
    # 写 data.js
    js_rows = []
    for r in merged:
        o = {}
        for k in CANON:
            v = r[k]
            if v == "" or v is None: o[k] = None
            else:
                try: o[k] = float(v) if ("." in str(v) or k in ("imp","clk","cost")) else (int(v) if str(v).isdigit() else v)
                except: o[k] = v
        if o.get("pid") is not None: o["pid"] = str(o["pid"])
        if o.get("sku") is not None: o["sku"] = str(o["sku"])
        js_rows.append(o)
    # 数据异常统计：商品维度缺款号 / 缺主图（按唯一商品ID去重，避免重复计数）
    anomaly = {"products_missing_sku":0, "products_missing_img":0,
               "total_products":0, "examples_sku":[], "examples_img":[]}
    prod_map = {}
    for r in merged:
        if r.get("dim") != "商品":
            continue
        pid = (r.get("pid") or "").strip()
        sku = (r.get("sku") or "").strip()
        img = (r.get("img") or "").strip()
        if pid in prod_map:
            o = prod_map[pid]
            if not o["sku"] and sku: o["sku"] = sku
            if not o["img"] and img: o["img"] = img
            continue
        prod_map[pid] = {"pid":pid, "sku":sku, "img":img}
    for o in prod_map.values():
        anomaly["total_products"] += 1
        if not o["sku"]:
            anomaly["products_missing_sku"] += 1
            if len(anomaly["examples_sku"]) < 8:
                anomaly["examples_sku"].append(o["pid"] or "(无商品ID)")
        if not o["img"]:
            anomaly["products_missing_img"] += 1
            if len(anomaly["examples_img"]) < 8:
                anomaly["examples_img"].append(o["sku"] or o["pid"] or "(无标识)")
    anomaly_js = ""
    if anomaly["total_products"]:
        anomaly_js = "window.DASHBOARD_ANOMALY = " + json.dumps(anomaly, ensure_ascii=False) + ";\n"
        print(f"  数据异常统计：商品 {anomaly['total_products']} 款，缺款号 {anomaly['products_missing_sku']}，缺主图 {anomaly['products_missing_img']}")

    # 合并预算配置（budget.json）进入 data.js，实现跨设备/浏览器永久保存
    budget_js = ""
    if os.path.exists(BUDGET_FILE):
        try:
            with open(BUDGET_FILE, encoding="utf-8") as bf:
                bcfg = json.load(bf)
            if isinstance(bcfg, dict) and (bcfg.get("total") or bcfg.get("plans")):
                budget_js = "window.DASHBOARD_BUDGET = " + json.dumps({
                    "total": float(bcfg.get("total") or 0),
                    "plans": bcfg.get("plans") or {}
                }, ensure_ascii=False) + ";\n"
                print(f"  已合并预算配置 budget.json：总预算 {bcfg.get('total',0)} 元，分计划 {len(bcfg.get('plans') or {})} 个")
        except Exception as e:
            print("  ⚠️ 读取 budget.json 失败，跳过预算合并：", e)
    with open(OUT_JS, "w", encoding="utf-8") as f:
        f.write(budget_js + anomaly_js + "window.DASHBOARD_DATA = " + json.dumps(js_rows, ensure_ascii=False) + ";\n")
    print(f"完成：合并 {len(merged)} 条 → master_data.csv + data.js")
    # 商品维度 款号 匹配统计
    prods = [r for r in merged if r.get("dim") == "商品"]
    if prods:
        hit = sum(1 for r in prods if r.get("sku"))
        print(f"  商品维度: {len(prods)} 个单品，其中 {hit} 个匹配到款号，{len(prods)-hit} 个未匹配(仅显示商品ID)")

if __name__ == "__main__":
    main()
